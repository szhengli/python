# -*- coding: utf-8 -*-
import openpyxl
import xlrd
import os

class ExcelHelper:
    def __init__(self, file_path, file_name):
        self.file_path = file_path
        self.file_name = file_name
        self.file = None
        self.workbook = None
        self.ws = None

    def file_abspath(self):
        if self.file_path.endswith('/'):
            self.file = self.file_path + self.file_name
        else:
            self.file = self.file_path + '/' + self.file_name
        # 判断文件是否存在，存在则返回
        if os.path.exists(self.file):
            return self.file
        else:
            exit("执行结束:导入文件不存在！")

    def add_sheet(self, sheet_name, load_file=None, create_sheet=True):
        if load_file is None:
            load_file = self.file

        # 打开创建一个新sheet
        if create_sheet:
            self.workbook = openpyxl.load_workbook(load_file)
            sheet_names = self.workbook.sheetnames
            if sheet_name in sheet_names:
                exit("已存在 '%s' 名称的sheet, 请重新选择名称。" % sheet_name)
            else:
                self.ws = self.workbook.create_sheet(sheet_name)

    def new_excel(self, sheet_name):
        self.workbook = openpyxl.Workbook()
        self.ws = self.workbook.active
        self.ws.title = sheet_name

    def get_excel_data(self, open_file=None, sheet_index=0, column_index=0):
        if open_file is None:
            open_file = self.file
        workbook = xlrd.open_workbook(open_file)
        ws = workbook.sheet_by_index(sheet_index)
        rows = ws.nrows
        data = []
        for i in range(column_index, rows):
            row_value = ws.row_values(i)
            row_value[-1] = str(xlrd.xldate_as_datetime(row_value[-1], 0))
            row_value[0] = int(row_value[0])
            data.append(str(tuple(row_value)))
        return data

    '''
    传入要修改的列宽：{'20':['A','B','D'],'15':['C']}
    ws.column_dimensions['A'].width=20
    '''
    def set_style(self, col):
        if isinstance(col, dict):
            for key in col:
                col_code = col[key]
                for i in range(len(col_code)):
                    set_col = col_code[i]
                    if '-' in set_col:
                        split_letter = set_col.split('-')
                        begin_letter, end_letter = split_letter
                        letter = [chr(x) for x in range(ord(begin_letter), ord(end_letter) + 1)]
                        for j in range(len(letter)):
                            set_col_split = letter[j]
                            self.ws.column_dimensions[set_col_split].width = key
                    else:
                        self.ws.column_dimensions[set_col].width = key

    def write_file(self, write_data):
        if write_data is None:
            exit('传入的值为空')

        # 准备写入数据:传入的必须是list,dict,tuple
        try:
            for line in range(len(write_data)):
                self.ws.append(write_data[line])
        except Exception as e:
            raise e

    def save_file(self):
        self.workbook.save(self.file)